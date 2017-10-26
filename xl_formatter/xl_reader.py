import re
from openpyxl import load_workbook, Workbook
from .models import TwitterData, InstagramData

def load_input(category_list, filename):
    file_path = 'xl_formatter/input_files/' + filename +'.xlsx'
    input_data = load_workbook(file_path, read_only=True)
    ws = input_data.active
    categories = {}

    for index, row in enumerate(ws.rows):

        if index == 0:
            for col in enumerate(row):
                category_index = col[0]
                category_title = col[1].value

                # add snippet to category_list
                if 'snippet' not in category_list:
                    category_list.append('snippet')

                # add pageType to category_list
                if 'pageType' not in category_list:
                    category_list.append('pageType')

                if category_title in category_list:
                    categories[category_title] = {}
                    categories[category_title]['index'] = category_index
            continue

        try:
            type = row[categories['pageType']['index']].value.lower()
        except:
            type = row[categories['pageType']['index']].value

        def extract_value(category):
            if category in category_list:
                category_value = row[categories[category]['index']].value
            else:
                return None

            return category_value

        def return_clout():
            clout_value = extract_value('clout')
            if clout_value == None or type(clout_value) is str:
                clout_value = 0

            return clout_value

        def return_content():
            content_value = extract_value('fulltext')
            if (content_value == 'None' or content_value == '' or content_value == None or content_value == 'NA'):
                content_value = row[categories['snippet']['index']].value

            return content_value

        def return_author():
            author_value = extract_value('author')
            try:
                author_value = author_value.lower()
                return author_value
            except:
                return author_value

        def return_followers():
            followers_value = extract_value('followers')

            if followers_value == None or followers_value == 'NA':
                followers_value = 0

            return int(followers_value)

        def return_brand_source():
            brand_source_value = extract_value('brand_source')
            try:
                brand_source_value = re.sub(r'\s', '', brand_source_value)
                return brand_source_value
            except:
                return brand_source_value


        if type == 'twitter':

            twitter_data = TwitterData(
                url=extract_value('url'),
                date=extract_value('date'),
                author_gender=extract_value('gender'),
                city=extract_value('city'),
                sentiment=extract_value('sentiment'),
                author_clout=return_clout(),
                contents=return_content(),
                author=return_author(),
                followers=return_followers(),
                brand_source=return_brand_source(),
                type=type
            )

            twitter_data.save()

        if type == 'instagram':

            instagram_data = InstagramData(
                url=extract_value('url'),
                date=extract_value('date'),
                author_gender=extract_value('gender'),
                city=extract_value('city'),
                sentiment=extract_value('sentiment'),
                author_clout=return_clout(),
                contents=return_content(),
                author=return_author(),
                followers=return_followers(),
                brand_source=return_brand_source(),
                type=type
            )

            instagram_data.save()

def split_input():
    print('Loading input data')
    file_path = 'xl_formatter/sun_care.xlsx'
    input_data = load_workbook(file_path, read_only=True)
    print('Data load complete!')
    print('Grabbing sheet')
    ws = input_data['Sheet']
    print('Grabbed that sheet!')

    print('start new workbook stuff')
    new_workbook = Workbook(write_only=True)
    new_sheet = new_workbook.create_sheet()
    print('workbook stuff done playa')

    counter = 0
    print('starting the writing process')
    for row in ws.iter_rows():
        row_data = []
        for cell in row:
            row_data.append(cell.value)
        print('appending row!', counter)
        counter += 1
        new_sheet.append(row_data)

    print('All Done writing playa!!!')
    print('Starting save!')
    new_workbook.save('xl_formatter/sun_care_sheet.xlsx')
    print('All done save!')




